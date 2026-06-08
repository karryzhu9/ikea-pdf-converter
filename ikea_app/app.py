import streamlit as st
import pdfplumber
import re
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

st.set_page_config(page_title="IKEA PDF → Excel", page_icon="🛋️", layout="wide")
st.title("🛋️ IKEA Planner PDF → Excel")
st.markdown("Upload your IKEA planner PDF and get a clean, two-sheet Excel file.")

# ──────────────────────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────────────────────
ARTICLE_RE  = re.compile(r'\b(\d{3}\.\d{3}\.\d{2})\b')                            # 602.143.31
CURRENCY_RE = re.compile(r'(?:€|EUR|£|\$)\s?(\d[\d.,]*\d|\d)')                    # €1,234.56
DECIMAL_RE  = re.compile(r'(?<![\d.,])(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?![\d])')  # 1,234.56 / 45.00
QTY_RE      = re.compile(r'(?<!\d)(\d+)\s*(?:x|pcs?|pieces|st)\b', re.IGNORECASE)
SECTION_RE  = re.compile(r'^\s*(\d+)\s*[\.\)]\s+(.{2,})$')                        # "1. Base cabinets"


def normalize_price(raw):
    """Handle 1,234.56 (US/UK) and 1.234,56 (EU) interchangeably."""
    s = (raw or "").strip().replace(' ', '')
    if not s:
        return None
    if ',' in s and '.' in s:
        s = (s.replace('.', '').replace(',', '.')
             if s.rfind(',') > s.rfind('.')
             else s.replace(',', ''))
    elif ',' in s:
        parts = s.split(',')
        s = s.replace(',', '.') if len(parts[-1]) == 2 else s.replace(',', '')
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _extract_prices(text):
    prices = [normalize_price(m.group(1)) for m in CURRENCY_RE.finditer(text)]
    prices = [p for p in prices if p is not None]
    if not prices:
        prices = [normalize_price(m.group(1)) for m in DECIMAL_RE.finditer(text)]
        prices = [p for p in prices if p is not None]
    return prices


def _starts_new_product(line):
    """IKEA series names begin with an ALL-CAPS token (METOD, KALLAX, MAXIMERA...)."""
    toks = line.strip().split()
    if not toks:
        return False
    first = toks[0]
    letters = [c for c in first if c.isalpha()]
    return len(first) >= 3 and len(letters) >= 3 and first == first.upper()


def _looks_like_name(s):
    letters = [c for c in s if c.isalpha()]
    if len(letters) < 2:
        return False
    return sum(c.isupper() for c in letters) / len(letters) > 0.6


def _block_to_item(block_lines, section):
    """Order-independent field extraction from one product block."""
    text = " ".join(block_lines)
    art_m = ARTICLE_RE.search(text)
    article_no = art_m.group(1) if art_m else ""

    text_wo_art = ARTICLE_RE.sub(' ', text)     # so article digits aren't read as a price
    prices = _extract_prices(text_wo_art)

    qty_m = QTY_RE.search(text_wo_art)
    qty = int(qty_m.group(1)) if qty_m else 1

    name, desc_parts = "", []
    for ln in block_lines:
        cleaned = ARTICLE_RE.sub('', ln)
        cleaned = CURRENCY_RE.sub('', cleaned)
        cleaned = DECIMAL_RE.sub('', cleaned)
        cleaned = QTY_RE.sub('', cleaned).strip(' .,-')
        if not cleaned:
            continue
        if not name and _looks_like_name(cleaned):
            name = cleaned
        else:
            desc_parts.append(cleaned)
    if not name and desc_parts:
        name = desc_parts.pop(0)

    if not name and not article_no:
        return None
    if not prices and not article_no:
        return None

    unit_price = total_price = None
    if len(prices) >= 2:
        lo, hi = min(prices), max(prices)
        if qty > 1 and abs(lo * qty - hi) <= 0.02 * max(hi, 1):
            unit_price, total_price = lo, hi
        else:
            unit_price, total_price = prices[0], prices[-1]
    elif len(prices) == 1:
        if qty > 1:
            unit_price = prices[0]
            total_price = round(unit_price * qty, 2)
        else:
            unit_price = total_price = prices[0]

    return {
        'Cabinet / Category': section,
        'Product Name': name,
        'Description': " ".join(desc_parts).strip(),
        'Article No.': article_no,
        'Qty': qty,
        'Unit Price (EUR)': unit_price if unit_price is not None else '',
        'Total Price (EUR)': total_price if total_price is not None else '',
        'Select (Y/N)': '',
    }


def parse_lines(all_lines):
    """Block-based parser — a product block runs until the next product begins."""
    items, section, buffer = [], "General", []

    def flush():
        if buffer:
            it = _block_to_item(buffer, section)
            if it:
                items.append(it)

    for raw in all_lines:
        line = raw.strip()
        if not line:
            continue

        sec = SECTION_RE.match(line)
        if sec and not ARTICLE_RE.search(line) and not _extract_prices(line):
            flush(); buffer.clear()
            section = line
            continue

        has_article = bool(ARTICLE_RE.search(line))
        buffer_has_article = any(ARTICLE_RE.search(b) for b in buffer)
        new_product = buffer and (
            (_starts_new_product(line) and buffer_has_article) or
            (has_article and buffer_has_article)
        )
        if new_product:
            flush(); buffer.clear()
        buffer.append(line)

    flush()
    return items


def parse_ikea_pdf(pdf_file):
    """Returns (items, raw_text). raw_text powers the debug view."""
    pages_text = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            pages_text.append(page.extract_text() or "")
    raw_text = "\n".join(pages_text)
    return parse_lines(raw_text.split("\n")), raw_text


# ──────────────────────────────────────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────────────────────────────────────
def autofit(ws, max_width=55):
    for col in ws.columns:
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            max(8, min(longest + 4, max_width))


def build_excel(items, selected_indices=None):
    selected_indices = selected_indices or []
    wb = Workbook()

    HDR = PatternFill("solid", fgColor="003087")
    ALT = PatternFill("solid", fgColor="EEF3FF")
    WHT = PatternFill("solid", fgColor="FFFFFF")
    CAT = PatternFill("solid", fgColor="D6E4FF")
    TOT = PatternFill("solid", fgColor="001F5E")
    thin = Side(style="thin", color="CCCCCC")
    BDR = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols = ['#', 'Cabinet / Category', 'Product Name', 'Description',
            'Article No.', 'Qty', 'Unit Price (EUR)', 'Total Price (EUR)']

    def header(ws, with_select):
        full = cols + (['Select (Y/N)'] if with_select else [])
        for ci, h in enumerate(full, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = HDR
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BDR
        ws.row_dimensions[1].height = 32

    def write_rows(ws, rows, with_select):
        for r, item in enumerate(rows, 2):
            shade = ALT if r % 2 == 0 else WHT
            ws.row_dimensions[r].height = 22
            n = ws.cell(row=r, column=1, value=r - 1)
            n.font = Font(name="Arial", size=10, bold=True, color="003087")
            n.fill = shade; n.border = BDR
            n.alignment = Alignment(horizontal="center", vertical="center")

            keys = cols[1:] + (['Select (Y/N)'] if with_select else [])
            for ci, key in enumerate(keys, 2):
                c = ws.cell(row=r, column=ci, value=item.get(key, ''))
                c.font = Font(name="Arial", size=10)
                c.fill = shade; c.border = BDR
                if ci in (7, 8):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cc = ws.cell(row=r, column=2)
            cc.font = Font(name="Arial", size=10, bold=True, color="051464")
            cc.fill = CAT

    def total_row(ws, n_rows, label, value, ncols):
        tr = n_rows + 2
        for ci in range(1, ncols + 1):
            ws.cell(row=tr, column=ci).fill = TOT
            ws.cell(row=tr, column=ci).border = BDR
        lbl = ws.cell(row=tr, column=7, value=label)
        lbl.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        lbl.alignment = Alignment(horizontal="right", vertical="center"); lbl.fill = TOT
        v = ws.cell(row=tr, column=8, value=value)
        v.font = Font(name="Arial", bold=True, color="FFDD00", size=11)
        v.number_format = '#,##0.00'
        v.alignment = Alignment(horizontal="right", vertical="center"); v.fill = TOT

    # Sheet 1 — All Items
    ws1 = wb.active
    ws1.title = "All Items"
    header(ws1, with_select=True)
    ws1.freeze_panes = "A2"
    write_rows(ws1, items, with_select=True)
    if items:
        n = len(items)
        dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
        dv.sqref = f"I2:I{n+1}"
        ws1.add_data_validation(dv)
        total_row(ws1, n, "GRAND TOTAL", f"=SUM(H2:H{n+1})", 9)
    autofit(ws1)

    # Sheet 2 — Selected Items
    ws2 = wb.create_sheet("Selected Items")
    header(ws2, with_select=False)
    selected = [items[i] for i in selected_indices]
    if selected:
        write_rows(ws2, selected, with_select=False)
        total_row(ws2, len(selected), "SELECTED TOTAL",
                  round(sum(items[i]['Total Price (EUR)'] or 0 for i in selected_indices), 2), 8)
    else:
        msg = ws2.cell(row=2, column=1,
                       value="No items selected — tick items in the app, or type Y in column I on 'All Items'.")
        msg.font = Font(name="Arial", size=10, italic=True, color="999999")
        ws2.merge_cells("A2:H2")
    autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload IKEA Planner PDF", type=["pdf"])

if uploaded:
    with st.spinner("Parsing PDF..."):
        items, raw_text = parse_ikea_pdf(uploaded)

    if not items:
        st.error("Couldn't parse any items. Open the debug view below to see the raw text.")
        with st.expander("🔍 Raw extracted text (debug)"):
            st.text(raw_text or "(no text extracted — the PDF may be scanned/image-only)")
    else:
        n_cats = len(set(i['Cabinet / Category'] for i in items))
        st.success(f"✅ Found **{len(items)} items** across **{n_cats} categories**")

        if 'selected_set' not in st.session_state:
            st.session_state.selected_set = set()

        st.subheader("Select items for your quotation")

        def cat_key(x):
            head = x.split('.')[0].strip()
            return int(head) if head.isdigit() else 999
        all_categories = sorted(set(i['Cabinet / Category'] for i in items), key=cat_key)

        selected_cats = st.multiselect(
            "Filter by category (leave blank to show all)",
            options=all_categories, default=[]
        )
        filtered = items if not selected_cats else \
            [i for i in items if i['Cabinet / Category'] in selected_cats]
        filtered_indices = [items.index(i) for i in filtered]

        display_df = pd.DataFrame([{
            'Select': idx in st.session_state.selected_set,
            'Category': i['Cabinet / Category'],
            'Product Name': i['Product Name'],
            'Description': i['Description'],
            'Article No.': i['Article No.'],
            'Qty': i['Qty'],
            'Unit Price': i['Unit Price (EUR)'],
            'Total': i['Total Price (EUR)'],
        } for idx, i in zip(filtered_indices, filtered)])

        edited = st.data_editor(
            display_df,
            column_config={
                "Select": st.column_config.CheckboxColumn("✅ Select", default=False, width="small"),
                "Category": st.column_config.TextColumn("Category", width="medium"),
                "Product Name": st.column_config.TextColumn("Product Name", width="medium"),
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Article No.": st.column_config.TextColumn("Article No.", width="small"),
                "Qty": st.column_config.NumberColumn("Qty", width="small"),
                "Unit Price": st.column_config.NumberColumn("Unit Price (€)", width="small", format="€%.2f"),
                "Total": st.column_config.NumberColumn("Total (€)", width="small", format="€%.2f"),
            },
            use_container_width=True, height=500, hide_index=True, key="editor",
        )

        # Reconcile selections directly from the editor's returned state (no stale closure).
        for pos, sel in enumerate(edited['Select'].tolist()):
            real_idx = filtered_indices[pos]
            if sel:
                st.session_state.selected_set.add(real_idx)
            else:
                st.session_state.selected_set.discard(real_idx)

        selected_indices = sorted(st.session_state.selected_set)
        selected_total = sum(items[i]['Total Price (EUR)'] or 0 for i in selected_indices)
        grand_total = sum(i['Total Price (EUR)'] or 0 for i in items)

        c1, c2 = st.columns(2)
        c1.metric("Grand Total", f"€{grand_total:,.2f}")
        c2.metric("Selected Total", f"€{selected_total:,.2f}",
                  delta=f"{len(selected_indices)} items selected")

        for i in selected_indices:
            items[i]['Select (Y/N)'] = 'Y'

        excel_buf = build_excel(items, selected_indices)
        st.download_button(
            "📥 Download Excel",
            data=excel_buf,
            file_name=uploaded.name.replace(".pdf", "_ItemList.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary",
        )

        with st.expander("🔍 Raw extracted text (debug — use this to verify parsing)"):
            st.text(raw_text)

st.markdown("---")
st.caption("IKEA Planner PDF → Excel · Sheet 1 = all items · Sheet 2 = your selection")
